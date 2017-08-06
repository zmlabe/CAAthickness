"""
Scripts reads in sea ice thickness data from CS-2 and PIOMAS for 
exploratory data analysis
 
Notes
-----
    Author : Zachary Labe
    Date   : 22 July 2017
"""

### Import modules
import numpy as np
from netCDF4 import Dataset
import matplotlib.pyplot as plt
import matplotlib.colors as c
import datetime
import read_SeaIceThick_PIOMAS as CT
import read_PiomasArea as CA
import statsmodels.api as sm
from mpl_toolkits.basemap import Basemap
from openpyxl import load_workbook
from scipy.interpolate import griddata as g
import statsmodels.api as sm
import scipy.stats as sts
import nclcmaps as ncm
import matplotlib.mlab as mlab
from matplotlib.ticker import NullFormatter

### Define directories
directorydata1 = '/surtsey/zlabe/seaice_obs/PIOMAS/'  
directorydata2 = '/surtsey/zlabe/seaice_obs/cs2/' 
directorydata3 = '/surtsey/zlabe/seaice_obs/PIOMAS/Thickness/'  
directoryfigure = '/home/zlabe/Desktop/'
#directoryfigure = '/home/zlabe/Documents/Research/SeaIceVariability/Figures/'

### Define time           
now = datetime.datetime.now()
currentmn = str(now.month)
currentdy = str(now.day)
currentyr = str(now.year)
currenttime = currentmn + '_' + currentdy + '_' + currentyr
titletime = currentmn + '/' + currentdy + '/' + currentyr
print '\n' '----Calc sea ice thickness data sets - %s----' % titletime 

### Alott time series
yearmin = 2010
yearmax = 2017
years = np.arange(yearmin,yearmax+1,1)
months = [r'April']

### Call functions
latsp,lonsp,sitp = CT.readPiomas(directorydata1,years,0.15)
area = CA.readPiomasArea(directorydata1)

#lonsp = lonsp-180.

datacs = np.genfromtxt(directorydata2 + 'test.txt',unpack=False,
                       usecols=[0,1,3],skip_header=1).tolist()
                                   
#datasort = np.asarray(sorted(datacs,key=operator.itemgetter(0)))   
datasort = np.asarray(datacs)

latcs = datasort[:,0]
loncs = datasort[:,1] 
sitcs = datasort[:,2]     

lon36 = np.where(loncs <= 0.)[0]
loncs[lon36] = loncs[lon36] + 360.
              
#datacs.sort(key=lambda x: x[0])      
#datacs.sort(key=lambda x: x[1])       
              
def piomasReader(directory,segment,years):
    
    filename = 'piomas_regrid1x1_sit_LENS_19792017.nc'
    
    data = Dataset(directory + filename)
    lat = data.variables['lat'][:,:]
    lon = data.variables['lon'][:,:]
    
#    lon = np.flipud(lon) - 180.
    
    if segment == 'sub':    # 1986-1994  
        timeslice = np.where((years >= 1986) & (years <= 1994))[0]
        sitp = data.variables['thick'][timeslice,:,:] 
    elif segment == 'icej':    # 2004-2009
        timeslice = np.where((years >= 2004) & (years <= 2009))[0]
        sitp = data.variables['thick'][timeslice,:,:]
    elif segment == 'cryo':    # 2011-2015
        timeslice = np.where((years >= 2011) & (years <= 2015))[0]
        sitp = data.variables['thick'][timeslice,:,:]
    else:
        sitp = data.variables['sit'][:,:,:,:]
    
    sitp[np.where(sitp == 0.)] = np.nan
        
    data.close()
    
    print 'Completed: PIOMAS data read!'
    return lat,lon,sitp
    
lat3,lon3,sit3 = piomasReader(directorydata3,None,years)
        
#### Read in other data - column 2 is warren snow
#wb = load_workbook(directorydata2 + 'cryosat_FYI.xlsx',read_only=True)
#ws = wb.get_sheet_by_name('cryosat')
## latitudes
#lats = np.array([r[0].value for r in ws.iter_rows()])
#latsq = np.asarray(map(lambda x: float(x),lats[1:]))
## longitudes
#lons = np.array([r[1].value for r in ws.iter_rows()])
#lonsq = np.asarray(map(lambdamask2 = np.isfinite(latmask) & np.isfinite(lonmask) x: float(x),lons[1:]))
## sea ice thickness from cryosat 
#sitc = np.array([r[3].value for r in ws.iter_rows()])
#sitcq = np.asarray(map(lambda x: float(x),sitc[1:]))

## Retrieve April SIT
sitpa = sit3[-1,3,:,:]
        
vals = g((lon3.flatten(),lat3.flatten()),sitpa.flatten(),(loncs,latcs))  
#f = si.Rbf(lat3,lon3,sitpa,kind='linear')
#z = f(latcs,loncs)

#############################################################################
#############################################################################
#############################################################################
#### Test Plot
varx = vals
vary = sitcs
timex = np.arange(vals.shape[0])
timey = np.arange(vals.shape[0])

fig = plt.figure()
plt.rc('text',usetex=True)
plt.rc('font',**{'family':'sans-serif','sans-serif':['Avant Garde']}) 

### Adjust axes in time series plots 
def adjust_spines(ax, spines):
    for loc, spine in ax.spines.items():
        if loc in spines:
            spine.set_position(('outward', 10))
        else:
            spine.set_color('none')  
    if 'left' in spines:
        ax.yaxis.set_ticks_position('left')
    else:
        ax.yaxis.set_ticks([])

    if 'bottom' in spines:
        ax.xaxis.set_ticks_position('bottom')
    else:
        ax.xaxis.set_ticks([]) 

### definitions for the axeats
left,width = 0.1,0.65
bottom,height = 0.1,0.65
bottom_h = left_h = left + width + 0.0

rect_scatter = [left,bottom,width,height]
rect_histx = [left,bottom_h,width,0.2]
rect_histy = [left_h,bottom,0.2,height]

axScatter = plt.axes(rect_scatter)
axHistx = plt.axes(rect_histx)
axHisty = plt.axes(rect_histy)

# no labels
nullfmt = NullFormatter()
axHistx.xaxis.set_major_formatter(nullfmt)
axHisty.yaxis.set_major_formatter(nullfmt)

smoothed = sm.nonparametric.lowess(vary,varx,it=0,frac=0.5)

mask = np.isfinite(varx) & np.isfinite(vary)
slope, intercept, r_value, p_value, std_err = sts.linregress(varx[mask],vary[mask])
line2 = slope*timex + intercept
        
axScatter.plot(timex,timey,color='k',linewidth=3,zorder=1)
axScatter.scatter(varx[mask],vary[mask],s=10,color='seagreen',edgecolor='darkgreen',
            linewidths=0.3,zorder=2,alpha=0.5)
axScatter.plot(smoothed[:,0],smoothed[:,1],color='m',linewidth=2,zorder=3)
#axScatter.plot(line2,linestyle='-',linewidth=2,color='m')

binwidth=0.25
bins = np.arange(0,5 + binwidth,binwidth)
n,bins,patches = axHistx.hist(varx[mask],bins=bins,normed=True,
             facecolor='dimgrey',edgecolor='w',alpha=1,linewidth=0.45)
n,binsy,patches = axHisty.hist(vary[mask],bins=bins,normed=True,
                               orientation='horizontal',
                               facecolor='dimgrey',edgecolor='w',
                               alpha=1,linewidth=0.45)

mu,sigma = sts.norm.fit(varx[mask])             
y = mlab.normpdf(bins,mu,sigma)
#lx = axHistx.plot(bins,y,'k-',linewidth=0.5)

mu2,sigma2 = sts.norm.fit(vary[mask])
y2 = mlab.normpdf(binsy,mu2,sigma2)
#ly = axHisty.plot(y2,binsy,'k-',linewidth=0.5)

axScatter.set_xlim((0,5))
axScatter.set_ylim((0,5))
axScatter.set_xticklabels(map(str,np.arange(0,6,1)))
axScatter.set_yticklabels(map(str,np.arange(0,6,1)))

axHistx.spines['top'].set_color('none')
axHistx.spines['right'].set_color('none')
axHistx.spines['left'].set_color('none')
axHistx.spines['bottom'].set_color('none')
axHisty.spines['top'].set_color('none')
axHisty.spines['right'].set_color('none')
axHisty.spines['left'].set_color('none')
axHisty.spines['bottom'].set_color('none')
axScatter.spines['top'].set_color('none')
axScatter.spines['right'].set_color('none')
axScatter.spines['left'].set_color('darkgrey')
axScatter.spines['bottom'].set_color('darkgrey')
axScatter.tick_params('both',length=4,width=1.5,which='major',color='darkgrey')

axHistx.xaxis.set_ticks([])
axHistx.yaxis.set_ticks([])
axHisty.xaxis.set_ticks([])
axHisty.yaxis.set_ticks([])

axHistx.set_yticklabels([])
axHisty.set_xticklabels([])

axScatter.xaxis.set_ticks_position('bottom')
axScatter.yaxis.set_ticks_position('left')

axScatter.grid(color='dimgrey',linewidth=0.4)

#### Add labels
axScatter.set_xlabel(r'\textbf{SIT( PIOMAS )(m)}')
axScatter.set_ylabel(r'\textbf{SIT( CS-2 )(m)}')

fig.subplots_adjust(top=0.95)
fig.subplots_adjust(bottom=0.15)

plt.savefig(directoryfigure + 'CAA/caatest2.png',dpi=300)

############################################################################
############################################################################
############################################################################
fig = plt.figure()
ax = plt.subplot(111)

plt.scatter(timex,varx,label=r'PIOMAS',color='m',zorder=2)
plt.scatter(timex,vary,label=r'CS-2',color='teal',zorder=1)

plt.legend()

plt.savefig(directoryfigure + 'CAA/caatest3.png',dpi=300)

############################################################################
############################################################################
############################################################################
fig = plt.figure()
ax = plt.subplot(111)

diff = varx - vary

plt.scatter(timex,diff,color='r',linewidth=1)

plt.savefig(directoryfigure + 'CAA/caatest4.png',dpi=300)
   
############################################################################
############################################################################
############################################################################
##### Define figure
latmask = latcs.copy()
lonmask = loncs.copy()   

latmask[np.where(np.isnan(varx))] = np.nan
lonmask[np.where(np.isnan(varx))] = np.nan
  
fig = plt.figure()
ax = plt.subplot(111)

style = 'polar'

if style == 'ortho':
    m = Basemap(projection='ortho',lon_0=-90,
                lat_0=70,resolution='l',round=True)
elif style == 'polar':
    m = Basemap(projection='npstere',boundinglat=55,lon_0=270,resolution='l',round =True)
    
m.drawmapboundary(fill_color='white')
m.drawcoastlines(color='darkgrey',linewidth=0.5)
parallels = np.arange(50,90,10)
meridians = np.arange(-180,180,30)
m.drawparallels(parallels,labels=[False,False,False,False],
                linewidth=0.3,color='k',fontsize=9)
m.drawmeridians(meridians,labels=[True,True,False,False],
                linewidth=0.3,color='k',fontsize=9)
m.drawlsmask(land_color='dimgrey',ocean_color='mintcream')

#### Plot filled contours  
cs = m.plot(lonmask,latmask,'yo',latlon=True)

fig.suptitle(r'\textbf{PIOMAS Interpolated Points}',color='dimgray',
          fontsize=18,alpha=1)
             
plt.subplots_adjust(top=0.85)


plt.savefig(directoryfigure + 'CAA/caatest5.png',dpi=300)

############################################################################
############################################################################
############################################################################
#### Define figure
fig = plt.figure()
ax = plt.subplot(111)

style = 'polar'

if style == 'ortho':
    m = Basemap(projection='ortho',lon_0=-90,
                lat_0=70,resolution='l',round=True)
elif style == 'polar':
    m = Basemap(projection='npstere',boundinglat=55,lon_0=270,resolution='l',round =True)
    
m.drawmapboundary(fill_color='white')
m.drawcoastlines(color='darkgrey',linewidth=0.5)
parallels = np.arange(50,90,10)
meridians = np.arange(-180,180,30)
m.drawparallels(parallels,labels=[False,False,False,False],
                linewidth=0.3,color='k',fontsize=9)
m.drawmeridians(meridians,labels=[True,True,False,False],
                linewidth=0.3,color='k',fontsize=9)
m.drawlsmask(land_color='dimgrey',ocean_color='mintcream')

#### Plot filled contours    
cs = m.contourf(lon3,lat3,sitpa,np.arange(0,5.1,0.1),latlon=True,
                extend='max')

barlim = np.arange(0,6,1) 
cs.set_cmap('plasma')
cbar = m.colorbar(cs,location='right',pad = 0.4,extend='max',
                  drawedges=False)
ticks = barlim
labels = map(str,barlim)
cbar.set_ticks(ticks)
cbar.set_ticklabels(labels)
cbar.set_label(r'\textbf{Thickness [m]}',fontsize=11,color='dimgray')
cbar.ax.tick_params(axis='y', size=.001)

plt.annotate(r'\textbf{April 2017 : PIOMAS}',xy=(0.1,1.1),
             xycoords='axes fraction',color='dimgray',fontsize=18,alpha=1) 

plt.subplots_adjust(top=0.85)

plt.savefig(directoryfigure + 'CAA/caatest6.png',dpi=300)

############################################################################
############################################################################
############################################################################
#### Define figure
fig = plt.figure()
ax = plt.subplot(111)

style = 'polar'

if style == 'ortho':
    m = Basemap(projection='ortho',lon_0=-90,
                lat_0=70,resolution='l',round=True)
elif style == 'polar':
    m = Basemap(projection='npstere',boundinglat=55,lon_0=270,resolution='l',round =True)
    
m.drawmapboundary(fill_color='white')
m.drawcoastlines(color='darkgrey',linewidth=0.5)
parallels = np.arange(50,90,10)
meridians = np.arange(-180,180,30)
m.drawparallels(parallels,labels=[False,False,False,False],
                linewidth=0.1,color='k',fontsize=5)
m.drawmeridians(meridians,labels=[True,True,False,False],
                linewidth=0.1,color='k',fontsize=5)
m.drawlsmask(land_color='dimgrey',ocean_color='mintcream')

#### Plot filled contours    
x,y = m(loncs,latcs)
diff[np.isnan(diff)]=0.0
cs = m.hexbin(x,y,C=diff,vmin = -2,vmax = 2)

barlim = np.arange(-2,3,1)
cmap = ncm.cmap('NCV_blu_red')  
cs.set_cmap(cmap)
cbar = m.colorbar(cs,location='right',pad = 0.4,extend='both',
                  drawedges=False)
ticks = barlim
labels = map(str,barlim)
cbar.set_ticks(ticks)
cbar.set_ticklabels(labels)
cbar.set_label(r'\textbf{Difference [m]}',fontsize=11,color='dimgray')
cbar.ax.tick_params(axis='y', size=.001)

plt.annotate(r'\textbf{April 2017 : [PIOMAS -- CS2]}',xy=(0.02,1.1),
             xycoords='axes fraction',color='dimgray',fontsize=18,alpha=1) 

plt.subplots_adjust(top=0.85)

plt.savefig(directoryfigure + 'CAA/caatest7.png',dpi=300)

###########################################################################
###########################################################################
###########################################################################
### Create text files
directorytext = '/home/zlabe/Documents/Projects/CAAthickness/Data/'

np.savetxt(directorytext + 'PIOMAS_vals_05_2017.txt',
           np.c_[vals,sitcs,latcs,loncs],
            fmt='%1.3f')